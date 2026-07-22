"""
Accept both .xlsx and .ods file uploads
Load data from either format using the existing load_spreadsheet_data() function
Immediately convert to a new Excel (.xlsx) workbook
Process the spreadsheet (add signature)
Download as .xlsx with proper MIME type
"""

import streamlit as st
import pandas as pd
import hashlib
import base64
import os
import tempfile
import zipfile
import io
import subprocess
import shutil
from datetime import datetime

from openpyxl import cell, load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from odf.opendocument import load as load_ods
from odf.table import Table, TableRow, TableCell
from odf.text import P
from odf import teletype

from cryptography.hazmat.primitives.asymmetric import rsa, padding
from cryptography.hazmat.primitives import serialization, hashes


import re

# =========================================================
# FIND APPROVER
# =========================================================

def find_approver(file_path):
    data = load_spreadsheet_data(file_path)

    pattern = r"^\s*(.*?)\s*:\s*(ok|approved)\s*$"

    for sheet_name, rows in data.items():

        if sheet_name == "Digital Signature":
            continue

        if not rows or not rows[0]:
            continue

        value = rows[0][0]

        if value is None:
            continue

        value = str(value)

        match = re.match(
            pattern,
            value,
            re.IGNORECASE
        )

        if match:
            approver_name = match.group(1).strip()
            return approver_name

    return None

# =========================================================
# HASH GENERATION
# =========================================================

def generate_hash_from_excel(file_path):
    data = load_spreadsheet_data(file_path)

    collected = []

    for sheet_name in sorted(data.keys()):
        if sheet_name == "Digital Signature":
            continue

        rows = data[sheet_name]
        collected.append(f"---SHEET:{sheet_name}---")

        for row in rows:
            if row is None:
                continue
            for value in row:
                if value is None:
                    value = ""

                value = str(value).strip()
                value = value.replace("\r", "")
                value = value.replace("\n", " ")

                try:
                    numeric = float(value)
                    if numeric.is_integer():
                        value = str(int(numeric))
                except:
                    pass

                collected.append(value)

    content = "|".join(collected)
    hash_value = hashlib.sha256(content.encode("utf-8")).hexdigest()
    return hash_value


# =========================================================
# ODS SUPPORT HELPERS
# =========================================================


def is_ods_file(file_path):
    return os.path.splitext(str(file_path))[1].lower() == ".ods"


def safe_get_attribute(elem, attr_name):
    if not hasattr(elem, "allowed_attributes"):
        return None
    try:
        allowed = elem.allowed_attributes()
        if allowed:
            allowed_args = [a[1].lower().replace("-", "") for a in allowed]
            if attr_name.lower().replace("-", "") not in allowed_args:
                return None
        return elem.getAttribute(attr_name)
    except Exception:
        return None


def sanitize_color(color_str):
    if not color_str:
        return None
    color_str = str(color_str).lstrip("#").strip().upper()
    if color_str in ("TRANSPARENT", "NONE", ""):
        return None
    if len(color_str) == 3:
        color_str = "".join(c * 2 for c in color_str)
    if len(color_str) in (6, 8) and all(c in "0123456789ABCDEF" for c in color_str):
        return color_str
    return None


def sanitize_sheet_name(name, existing_names):
    if not name:
        name = "Sheet"
    for ch in [":", "\\", "/", "?", "*", "[", "]"]:
        name = name.replace(ch, "_")
    name = name[:31].strip()
    if not name:
        name = "Sheet"

    base_name = name
    counter = 1
    while name in existing_names:
        suffix = f"_{counter}"
        name = base_name[: 31 - len(suffix)] + suffix
        counter += 1

    existing_names.add(name)
    return name


def get_ods_cell_text(cell):
    try:
        text = teletype.extractText(cell)
        if text is not None:
            text = text.strip()
        if text:
            return text
    except Exception:
        pass

    for attr in ["value", "stringvalue", "datevalue", "booleanvalue"]:
        val = safe_get_attribute(cell, attr)
        if val is not None and str(val).strip() != "":
            return str(val).strip()

    return ""


def load_ods_data(file_path):
    doc = load_ods(file_path)
    data = {}

    for table in doc.spreadsheet.getElementsByType(Table):
        sheet_name = safe_get_attribute(table, "name") or "Sheet"
        rows = []

        for row in table.getElementsByType(TableRow):
            values = []
            if hasattr(row, "childNodes"):
                for child in row.childNodes:
                    tag = getattr(child, "tagName", "")
                    if tag in ("table:table-cell", "table:covered-table-cell"):
                        repeat_str = safe_get_attribute(child, "numbercolumnsrepeated")
                        repeat = int(repeat_str) if repeat_str else 1
                        cell_value = get_ods_cell_text(child) if tag == "table:table-cell" else ""
                        for _ in range(repeat):
                            values.append(cell_value)

            while values and values[-1] == "":
                values.pop()

            row_repeat_str = safe_get_attribute(row, "numberrowsrepeated")
            row_repeat = int(row_repeat_str) if row_repeat_str else 1

            for _ in range(row_repeat):
                rows.append(values.copy())

        data[sheet_name] = rows

    return data


def _convert_with_excel(input_path, output_path):
    import win32com.client

    excel = None
    workbook = None

    try:
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False

        workbook = excel.Workbooks.Open(
            os.path.abspath(input_path),
            ReadOnly=True
        )

        workbook.SaveAs(
            os.path.abspath(output_path),
            FileFormat=51
        )
    finally:
        if workbook is not None:
            try:
                workbook.Close(SaveChanges=False)
            except Exception:
                pass
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass


def _find_libreoffice_command():
    candidates = [
        shutil.which("soffice"),
        shutil.which("libreoffice"),
        r"C:\Program Files\LibreOffice\program\soffice.exe",
        r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
        r"C:\Program Files\OpenOffice\program\soffice.exe",
        r"C:\Program Files (x86)\OpenOffice\program\soffice.exe"
    ]

    for candidate in candidates:
        if candidate and os.path.exists(candidate):
            return candidate

    return None


def _convert_with_libreoffice(input_path, output_path):
    libreoffice_cmd = _find_libreoffice_command()

    if not libreoffice_cmd:
        raise RuntimeError("LibreOffice is not available on this system.")

    command = [
        libreoffice_cmd,
        "--headless",
        "--convert-to",
        "xlsx",
        "--outdir",
        os.path.dirname(os.path.abspath(output_path)),
        os.path.abspath(input_path)
    ]

    result = subprocess.run(
        command,
        capture_output=True,
        text=True,
        check=False
    )

    if result.returncode != 0:
        raise RuntimeError(
            f"LibreOffice conversion failed: {result.stderr.strip() or result.stdout.strip()}"
        )

    converted_name = os.path.splitext(os.path.basename(input_path))[0] + ".xlsx"
    converted_path = os.path.join(
        os.path.dirname(os.path.abspath(output_path)),
        converted_name
    )

    if not os.path.exists(converted_path):
        raise RuntimeError("LibreOffice conversion did not produce an XLSX file.")

    os.replace(converted_path, output_path)


def convert_ods_to_xlsx(input_path, output_path):
    try:
        _convert_with_excel(input_path, output_path)
        return
    except Exception:
        pass

    try:
        _convert_with_libreoffice(input_path, output_path)
        return
    except Exception:
        pass

    try:
        _convert_with_python(input_path, output_path)
        return
    except Exception as exc:
        raise RuntimeError(
            f"Unable to convert ODS file to XLSX: {exc}"
        ) from exc


def format_worksheet_layout(ws):
    thin_border = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0")
    )

    # First pass: set column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value or "")
            if "\n" in val:
                lines = val.split("\n")
                line_len = max(len(l) for l in lines) if lines else 0
            else:
                line_len = len(val)
            if line_len > max_len:
                max_len = line_len
        if max_len > 0:
            calculated_w = max(max_len + 6, 18)
            ws.column_dimensions[col_letter].width = min(calculated_w, 55)

    # Second pass: set row heights, alignment, borders
    for row in ws.iter_rows():
        if not any(cell.value is not None and str(cell.value).strip() != "" for cell in row):
            continue

        row_num = row[0].row

        # Calculate a good row height based on longest cell content
        max_lines = 1
        for cell in row:
            if cell.value is not None:
                val = str(cell.value)
                col_letter = get_column_letter(cell.column)
                col_width = ws.column_dimensions[col_letter].width or 18
                char_per_line = max(int(col_width * 1.2), 10)
                wrapped_lines = max(
                    len(val) // char_per_line + 1,
                    val.count("\n") + 1
                )
                if wrapped_lines > max_lines:
                    max_lines = wrapped_lines

        row_height = max(24, min(max_lines * 18, 120))
        ws.row_dimensions[row_num].height = row_height

        for cell in row:
            if cell.value is not None and str(cell.value).strip() != "":
                existing_h = cell.alignment.horizontal if cell.alignment else None
                cell.alignment = Alignment(
                    horizontal=existing_h if existing_h and existing_h != "general" else "left",
                    vertical="center",
                    wrap_text=True
                )
                if not cell.border or not cell.border.left.style:
                    cell.border = thin_border



def parse_odf_styles(doc):
    styles = {}
    sources = []
    if hasattr(doc, "automaticstyles") and doc.automaticstyles:
        sources.extend(list(doc.automaticstyles.childNodes))
    if hasattr(doc, "styles") and doc.styles:
        sources.extend(list(doc.styles.childNodes))

    for s in sources:
        name = safe_get_attribute(s, "name")
        if not name:
            continue
        style_info = {
            "bg_color": None,
            "bold": False,
            "align": None,
            "color": None,
            "border": False
        }
        if hasattr(s, "childNodes"):
            for child in s.childNodes:
                if not hasattr(child, "attributes"):
                    continue
                for (ns, attr), val in child.attributes.items():
                    if attr == "background-color" and val != "transparent":
                        style_info["bg_color"] = val.lstrip("#").upper()
                    elif attr == "font-weight" and val in ("bold", "700", "800", "900"):
                        style_info["bold"] = True
                    elif attr == "text-align":
                        style_info["align"] = val
                    elif attr == "color":
                        style_info["color"] = val.lstrip("#").upper()
                    elif "border" in attr and val != "none":
                        style_info["border"] = True
        styles[name] = style_info
    return styles


def _convert_with_python(input_path, output_path):
    doc = load_ods(input_path)
    styles = parse_odf_styles(doc)

    wb = Workbook()
    existing_sheet_names = set()
    first_sheet = True

    thin_border = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0")
    )

    for table in doc.spreadsheet.getElementsByType(Table):
        raw_name = safe_get_attribute(table, "name")
        sheet_name = sanitize_sheet_name(raw_name, existing_sheet_names)
        if first_sheet:
            ws = wb.active
            ws.title = sheet_name
            first_sheet = False
        else:
            ws = wb.create_sheet(sheet_name)

        row_idx = 1
        for row in table.getElementsByType(TableRow):
            row_repeat_str = safe_get_attribute(row, "numberrowsrepeated")
            row_repeat = int(row_repeat_str) if row_repeat_str else 1

            cells_data = []
            has_row_data = False

            if hasattr(row, "childNodes"):
                for child in row.childNodes:
                    tag = getattr(child, "tagName", "")
                    if tag in ("table:table-cell", "table:covered-table-cell"):
                        repeat_str = safe_get_attribute(child, "numbercolumnsrepeated")
                        repeat = int(repeat_str) if repeat_str else 1

                        if tag == "table:table-cell":
                            cell_value = get_ods_cell_text(child)
                            style_name = safe_get_attribute(child, "stylename")
                            style_info = styles.get(style_name, {}) if style_name else {}
                        else:
                            cell_value = ""
                            style_info = {}

                        for _ in range(repeat):
                            cells_data.append((cell_value, style_info))
                            if cell_value != "":
                                has_row_data = True

            while cells_data and cells_data[-1][0] == "":
                cells_data.pop()

            if not has_row_data and row_idx > 1 and row_repeat > 5:
                break

            for _ in range(row_repeat):
                for col_idx, (val, st_info) in enumerate(cells_data, 1):
                    c = ws.cell(row=row_idx, column=col_idx)
                    c.value = val

                    try:
                        bg = sanitize_color(st_info.get("bg_color"))
                        if bg:
                            c.fill = PatternFill(
                                start_color=bg,
                                end_color=bg,
                                fill_type="solid"
                            )

                        is_bold = st_info.get("bold", False)
                        font_color = sanitize_color(st_info.get("color"))
                        if is_bold or font_color:
                            c.font = Font(
                                bold=is_bold,
                                color=font_color if font_color else "000000"
                            )

                        align = st_info.get("align")
                        if align in ("left", "center", "right", "justify"):
                            c.alignment = Alignment(horizontal=align)

                        if st_info.get("border"):
                            c.border = thin_border
                    except Exception:
                        pass

                row_idx += 1

        try:
            format_worksheet_layout(ws)
        except Exception:
            pass

    if len(wb.sheetnames) == 0:
        wb.create_sheet("Sheet1")

    wb.save(output_path)



def load_spreadsheet_data(file_path):
    if is_ods_file(file_path):
        return load_ods_data(file_path)

    wb = load_workbook(file_path, data_only=False)
    data = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []

        for row in ws.iter_rows():
            rows.append([cell.value for cell in row])

        data[sheet_name] = rows

    return data


# =========================================================
# GENERATE NEW KEY PAIR
# =========================================================

def generate_new_keys():

    private_key = rsa.generate_private_key(
        public_exponent=65537,
        key_size=2048
    )

    public_key = private_key.public_key()

    private_key_pem = private_key.private_bytes(
        encoding=serialization.Encoding.PEM,
        format=serialization.PrivateFormat.PKCS8,
        encryption_algorithm=serialization.NoEncryption()
    )

    public_key_pem = public_key.public_bytes(
        encoding=serialization.Encoding.PEM,
        format=serialization.PublicFormat.SubjectPublicKeyInfo
    )

    return private_key_pem, public_key_pem



# =========================================================
# SIGN HASH
# =========================================================

def sign_hash(hash_value, private_key):

    signature = private_key.sign(

        hash_value.encode(),

        padding.PKCS1v15(),

        hashes.SHA256()
    )

    signature_b64 = base64.b64encode(signature).decode()

    return signature_b64


# =========================================================
# STORE SIGNATURE IN EXCEL
# =========================================================

def store_signature(
    file_path,
    signature_b64,
    approver_name
):
    # File is already in .xlsx format at this point
    wb = load_workbook(file_path)

    if "Digital Signature" in wb.sheetnames:
        del wb["Digital Signature"]

    ws = wb.create_sheet("Digital Signature")

    crc_value = hashlib.sha256(
        signature_b64.encode()
    ).hexdigest()[:16]

    ws["A1"] = crc_value
    ws["A2"] = signature_b64

    current_date = datetime.now().strftime(
        "%d-%m-%Y"
    )

    ws["A3"] = (
        f"Approved by {approver_name} "
        f"on {current_date}"
    )

    for sheet in wb.worksheets:
        format_worksheet_layout(sheet)

    wb.save(file_path)





# =========================================================
# STREAMLIT UI
# =========================================================

st.set_page_config(
    page_title="Digital Signature Generator",
    page_icon="✍️",
    layout="centered"
)

st.title("Digital Signature Generator")

uploaded_file = st.file_uploader(
    "Upload Spreadsheet File",
    type=["xlsx", "ods"]
)


uploaded_private_key = st.file_uploader(
    "Upload Private Key (.pem)",
    type=["pem"]
)

generate_new = st.button("Generate And Download New Keys")


# =========================================================
# GENERATE NEW KEY
# =========================================================

if generate_new:

    private_key_pem, public_key_pem = generate_new_keys()

    public_key_text = public_key_pem.decode()

    st.success("New key pair generated")

    st.subheader("PUBLIC KEY")

    st.text_area(
        "Share this public key with receiver",
        public_key_text,
        height=250
    )

    # =========================================
    # CREATE ZIP IN MEMORY
    # =========================================

    zip_buffer = io.BytesIO()

    with zipfile.ZipFile(
        zip_buffer,
        "w",
        zipfile.ZIP_DEFLATED
    ) as zip_file:

        zip_file.writestr(
            "private_key.pem",
            private_key_pem
        )

        zip_file.writestr(
            "public_key.pem",
            public_key_pem
        )

    zip_buffer.seek(0)

    # =========================================
    # SINGLE DOWNLOAD BUTTON
    # =========================================

    st.download_button(
        label="Download Keys ZIP",
        data=zip_buffer,
        file_name="RSA_Keys.zip",
        mime="application/zip"
    )



if uploaded_file is not None and uploaded_private_key is not None:

    file_ext = os.path.splitext(uploaded_file.name)[1].lower()
    if file_ext not in [".xlsx", ".ods"]:
        file_ext = ".xlsx"

    # Create temp file with original extension to load the file
    with tempfile.NamedTemporaryFile(
        delete=False,
        suffix=file_ext
    ) as tmp:
        tmp.write(uploaded_file.read())
        temp_file_path = tmp.name

    if file_ext == ".ods":
        with tempfile.NamedTemporaryFile(
            delete=False,
            suffix=".xlsx"
        ) as tmp:
            temp_path = tmp.name

        try:
            convert_ods_to_xlsx(temp_file_path, temp_path)
        except RuntimeError as exc:
            try:
                os.remove(temp_file_path)
            except Exception:
                pass
            st.error(str(exc))
            st.stop()
        finally:
            try:
                os.remove(temp_file_path)
            except Exception:
                pass
    else:
        temp_path = temp_file_path

    try:

        # =========================================
        # CHECK APPROVAL
        # =========================================

        approver_name = find_approver(
            temp_path
        )

        if approver_name is None:

            st.error(
                "Approval not found. "
                "A1 of at least one sheet must contain "
                "'Name : OK' or 'Name : Approved'"
            )

            st.stop()



        # Load uploaded private key
        private_key = serialization.load_pem_private_key(
            uploaded_private_key.read(),
            password=None
        )

        # Generate hash
        hash_value = generate_hash_from_excel(
            temp_path
        )

        # Sign hash
        signature_b64 = sign_hash(
            hash_value,
            private_key
        )

        # Store signature
        store_signature(
            temp_path,
            signature_b64,
            approver_name
        )

        st.success(
            "Digital signature stored successfully"
        )

        with open(temp_path, "rb") as f:
            base, ext = os.path.splitext(uploaded_file.name)
            # Always download as .xlsx
            st.download_button(
                label="Download Signed Spreadsheet",
                data=f,
                file_name=f"{base}_Hashed.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    except Exception as e:

        st.error(str(e))
